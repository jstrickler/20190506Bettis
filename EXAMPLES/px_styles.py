#!/usr/bin/env python
# (c)2015 John Strickler
import openpyxl as px

print(px.__version__)

def main():
    """program entry point"""
    wb = px.load_workbook('../DATA/presidents.xlsx')
    ws = wb.get_sheet_by_name('US Presidents')

    update_last_names(ws)

    wb.save('presidents3.xlsx')

def update_last_names(ws):
    """Make the last name column blue and bold"""
    for row in ws['B2:B45']:
        cell = row[0]
        cell.value = cell.value.upper()
        cell.font = px.styles.Font(color='99000099', name="Comic Sans")
        cell.fill = px.styles.PatternFill(patternType='solid', bgColor=px.styles.Color("2000AA00"))

if __name__ == '__main__':
    main()
